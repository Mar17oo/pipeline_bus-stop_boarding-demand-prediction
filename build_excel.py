"""Build supervisor comparison Excel workbook."""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

cv   = pd.read_csv('results_cv.csv')
summ = pd.read_csv('results_summary.csv')

wb = openpyxl.Workbook()

HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
SUB_FILL  = PatternFill('solid', fgColor='2E75B6')
ALT_FILL  = PatternFill('solid', fgColor='D9E1F2')
WIN_FILL  = PatternFill('solid', fgColor='E2EFDA')
ZHN_FILL  = PatternFill('solid', fgColor='FFF2CC')
RED_FILL  = PatternFill('solid', fgColor='FFE0E0')
HDR_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
SUB_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
REG_FONT  = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', bold=True, size=10)
thin = Side(style='thin', color='999999')
REG_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def hdr(ws, row, col, val, fill=None, font=None, align=CTR):
    fill = fill or HDR_FILL
    font = font or HDR_FONT
    c = ws.cell(row, col, val)
    c.fill = fill
    c.font = font
    c.alignment = align
    c.border = REG_BORDER


def cell(ws, row, col, val, bold=False, fill=None, align=CTR):
    c = ws.cell(row, col, val)
    c.font = BOLD_FONT if bold else REG_FONT
    c.alignment = align
    c.border = REG_BORDER
    if fill:
        c.fill = fill


# =====================================================================
# SHEET 1 — Summary
# =====================================================================
ws = wb.active
ws.title = '1. Results Summary'
ws.sheet_view.showGridLines = False

ws.merge_cells('A1:J1')
t = ws.cell(1, 1, 'MSc Dissertation – GATv2 Cold-Start Bus Demand | Results vs Zheng et al. (2025)')
t.font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
t.alignment = CTR
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:J2')
sub = ws.cell(2, 1, 'University of Essex | CE902 | Leave-Borough-Out Spatial CV (33 London Boroughs) | Epochs = 1,000')
sub.font = Font(name='Calibri', size=10, color='595959', italic=True)
sub.alignment = CTR
ws.row_dimensions[2].height = 18

# --- YOUR results ---
ws.merge_cells('A4:J4')
hdr(ws, 4, 1, 'YOUR MODEL — London TfL BUSTO, 17,943 Stops, Inductive Cold-Start, 33-Fold Leave-Borough-Out CV')
ws.row_dimensions[4].height = 22

cols = ['Model', 'WMAPE Mean', 'WMAPE Std', 'WMAPE Median',
        'RMSE Mean', 'RMSE Std', 'RMSE Median',
        'MAE Mean', 'MAE Std', 'MAE Median']
for j, c in enumerate(cols, 1):
    hdr(ws, 5, j, c, fill=SUB_FILL)
ws.row_dimensions[5].height = 22

labels = {
    'HistAvg': 'Historical Average (Baseline)',
    'RF':      'Random Forest (Baseline)',
    'MLP':     'MLP Neural Net (Ablation)',
    'GATv2':   'GATv2 Multigraph (Proposed)',
}
for i, row in summ.iterrows():
    r = i + 6
    g = row['model'] == 'GATv2'
    f = WIN_FILL if g else None
    cell(ws, r, 1,  labels[row['model']],               bold=g, fill=f, align=LFT)
    cell(ws, r, 2,  round(row['WMAPE_mean'], 4),         bold=g, fill=f)
    cell(ws, r, 3,  f"+/-{round(row['WMAPE_std'], 4)}", bold=g, fill=f)
    cell(ws, r, 4,  round(row['WMAPE_median'], 4),       bold=g, fill=f)
    cell(ws, r, 5,  round(row['RMSE_mean'], 1),          bold=g, fill=f)
    cell(ws, r, 6,  f"+/-{round(row['RMSE_std'], 1)}",  bold=g, fill=f)
    cell(ws, r, 7,  round(row['RMSE_median'], 1),        bold=g, fill=f)
    cell(ws, r, 8,  round(row['MAE_mean'], 1),           bold=g, fill=f)
    cell(ws, r, 9,  f"+/-{round(row['MAE_std'], 1)}",   bold=g, fill=f)
    cell(ws, r, 10, round(row['MAE_median'], 1),         bold=g, fill=f)

# --- ZHENG results ---
ws.merge_cells('A11:J11')
hdr(ws, 11, 1, 'ZHENG ET AL. (2025) MF-STGAT — Beijing IC Card, 510 Stops, Transductive, 7:3 Split, 1,000 Epochs, 15-min prediction')
ws.row_dimensions[11].height = 22

for j, c in enumerate(cols, 1):
    hdr(ws, 12, j, c, fill=SUB_FILL)
ws.row_dimensions[12].height = 22

zheng = [
    ('HA Baseline',                0.5320, '-', '-', 37.22, '-', '-', 22.41, '-', '-'),
    ('ARIMA Baseline',             0.3958, '-', '-', 28.14, '-', '-', 17.29, '-', '-'),
    ('LSTM Baseline',              0.2185, '-', '-', 15.74, '-', '-',  8.64, '-', '-'),
    ('GCN-LSTM Baseline',          0.2012, '-', '-', 14.43, '-', '-',  7.98, '-', '-'),
    ('MF_STGAT Proposed @15-min', 0.1360, '-', '-',  1.49, '-', '-',  1.04, '-', '-'),
]
for i, (label, *vals) in enumerate(zheng):
    r = i + 13
    best = 'MF_STGAT' in label
    f = ZHN_FILL if best else None
    cell(ws, r, 1, label, bold=best, fill=f, align=LFT)
    for j, v in enumerate(vals, 2):
        cell(ws, r, j, v, bold=best, fill=f)

# --- Note ---
ws.merge_cells('A19:J20')
note = ws.cell(19, 1,
    'COMPARABILITY NOTE: Metrics are NOT directly comparable between studies. '
    'Zheng et al. predict the next 15-min interval flow for SEEN stops using historical sequences (transductive). '
    'This work predicts total daily demand for UNSEEN boroughs with zero historical data (inductive cold-start). '
    'The valid within-study comparison: GATv2 WMAPE 0.8381 vs HistAvg 1.0822 = 23% improvement. '
    'Zheng WMAPE 0.136 is for a different (easier) task and cannot be numerically compared.')
note.font = Font(name='Calibri', size=9, color='7F0000', italic=True)
note.alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[19].height = 35
ws.row_dimensions[20].height = 25

# --- Methodology table ---
ws.merge_cells('A22:J22')
hdr(ws, 22, 1, 'METHODOLOGICAL COMPARISON — Parameters That CAN Be Compared')
ws.row_dimensions[22].height = 22

meth_hdrs = ['Parameter', 'This Work (GATv2)', 'Zheng et al. (2025) MF-STGAT', 'Comparable?', '']
for j, c in enumerate(meth_hdrs[:4], 1):
    hdr(ws, 23, j, c, fill=SUB_FILL)

meth = [
    ('GNN Type',           'GATv2 (Brody et al. 2022)',              'GAT (Velickovic et al. 2018)',              'Similar YES'),
    ('Hidden Dimension',   '64',                                      '64',                                        'MATCH YES'),
    ('Training Epochs',    '1,000 + early stopping (patience=40)',   '1,000 + early stopping',                    'MATCH YES'),
    ('Multigraph Edges',   'KNN geographic + route connectivity',    'KNN geographic + POI functional similarity','Similar YES'),
    ('Spatial Features',   '8 ONS accessibility + lat/lon (10 total)','POI buffer 500m (5 categories)',           'Equivalent YES'),
    ('Metrics',            'MAE, RMSE, WMAPE',                       'MAE, RMSE, WMAPE',                          'MATCH YES'),
    ('Prediction target',  'Total daily boardings (static)',         '15-min interval flow (temporal)',           'DIFFERENT NO'),
    ('Setting',            'Inductive - entire boroughs unseen',     'Transductive - all stops seen in training','DIFFERENT NO'),
    ('Dataset city',       'London, UK (open TfL data)',             'Beijing, China (proprietary IC card)',      'DIFFERENT NO'),
    ('Network scale',      '17,943 stops, 33 boroughs (larger)',     '510 stops, 12 bus lines',                  'Different scale'),
]
for i, (param, ours, zheng_val, match) in enumerate(meth):
    r = i + 24
    ok = 'YES' in match and 'DIFFERENT' not in match
    bad = 'DIFFERENT' in match
    fill = WIN_FILL if ok else (RED_FILL if bad else None)
    cell(ws, r, 1, param,      bold=True, align=LFT)
    cell(ws, r, 2, ours,       align=LFT)
    cell(ws, r, 3, zheng_val,  align=LFT)
    cell(ws, r, 4, match,      bold=ok, fill=fill)

ws.column_dimensions['A'].width = 28
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws.column_dimensions[col].width = 16

# =====================================================================
# SHEET 2 — Per-Borough
# =====================================================================
ws2 = wb.create_sheet('2. Per-Borough Results')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:G1')
t2 = ws2.cell(1, 1, 'Per-Borough WMAPE — 33-Fold Leave-Borough-Out Spatial CV (green = GATv2 wins)')
t2.font = Font(name='Calibri', bold=True, size=13, color='1F4E79')
t2.alignment = CTR
ws2.row_dimensions[1].height = 26

for j, label in enumerate(['Borough', 'Test Stops', 'HistAvg', 'RF', 'MLP', 'GATv2 (Proposed)', 'Best Model'], 1):
    hdr(ws2, 2, j, label)
ws2.row_dimensions[2].height = 22

pivot = cv.pivot_table(index=['borough', 'n_test'], columns='model', values='WMAPE').reset_index()
pivot.columns.name = None
pivot = pivot.sort_values('borough').reset_index(drop=True)

models_ordered = ['HistAvg', 'RF', 'MLP', 'GATv2']
for i, row in pivot.iterrows():
    vals = [row.get(m, float('nan')) for m in models_ordered]
    best_v = min(vals)
    best_m = models_ordered[vals.index(best_v)]
    gatv2_wins = best_m == 'GATv2'
    row_f = WIN_FILL if gatv2_wins else (ALT_FILL if i % 2 == 0 else None)

    cell(ws2, i + 3, 1, row['borough'],    align=LFT, fill=row_f)
    cell(ws2, i + 3, 2, int(row['n_test']), fill=row_f)
    for j, (m, v) in enumerate(zip(models_ordered, vals), 3):
        is_best_col = v == best_v
        cell(ws2, i + 3, j, round(v, 4),
             bold=is_best_col,
             fill=WIN_FILL if is_best_col and gatv2_wins else row_f)
    cell(ws2, i + 3, 7,
         f"GATv2 ★" if gatv2_wins else best_m,
         bold=gatv2_wins, fill=row_f)

ws2.column_dimensions['A'].width = 28
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws2.column_dimensions[col].width = 18

# =====================================================================
# SHEET 3 — Email Template
# =====================================================================
ws3 = wb.create_sheet('3. Email to Zheng Authors')
ws3.sheet_view.showGridLines = False

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 90

fields = [
    ('To:', 'Prof. Wenquan Li (corresponding author) — wenqli@seu.edu.cn'),
    ('CC:', 'Yan Zheng — 230228862@seu.edu.cn'),
    ('Institution:', 'School of Transportation, Southeast University, Nanjing, Jiangsu 211189, PR China'),
    ('Subject:', 'Research Data Request — Beijing Bus IC Card Dataset (MF-STGAT, DOI: 10.1061/JTEPBS.TEENG-8794)'),
]
for i, (k, v) in enumerate(fields, 1):
    ws3.cell(i, 1, k).font = Font(bold=True, size=11)
    ws3.cell(i, 2, v).font = Font(size=11)
    ws3.row_dimensions[i].height = 18

body = [
    '',
    'Dear Professor Li and co-authors,',
    '',
    'I am Maria Isabel Bautista Hernandez, an MSc Computer Engineering student at the University of',
    'Essex (UK), supervised by Dr. Vishal K. Singh. I am completing my dissertation on inductive',
    'Graph Attention Network (GATv2) for cold-start bus stop demand prediction in London.',
    '',
    'I read your paper "Predicting Regional-Level Bus Stop Passenger Flow with a Multigraph Fusion',
    'Spatio-Temporal Graph Attention Network" (JTEPBS.TEENG-8794, Journal of Transportation',
    'Engineering Part A, 2025) and found it directly relevant to my research.',
    '',
    'My work specifically addresses the cold-start gap mentioned in your future work section:',
    'predicting demand for bus stops with NO prior historical passenger flow data (inductive setting).',
    'I apply GATv2 with a leave-borough-out cross-validation across 33 London boroughs.',
    '',
    'I would greatly appreciate the opportunity to test my framework on your Beijing IC card dataset',
    '(12 bus lines, 510 stops, Dongcheng/Xicheng, April 2019) to provide a direct comparison.',
    'I am fully willing to sign a data sharing/non-disclosure agreement and will use the data',
    'exclusively for academic research, citing your paper fully.',
    '',
    'Would it be possible to share the dataset under appropriate restrictions?',
    '',
    'Thank you very much for your excellent work and for considering this request.',
    '',
    'Kind regards,',
    'Maria Isabel Bautista Hernandez',
    'MSc Computer Engineering, University of Essex, UK',
    'Email: arturomar0210@gmail.com',
]
for i, line in enumerate(body, len(fields) + 2):
    ws3.cell(i, 2, line).font = Font(size=11)
    ws3.row_dimensions[i].height = 16

wb.save('supervisor_comparison_results.xlsx')
print('Saved: supervisor_comparison_results.xlsx')
print('Sheets: 1. Results Summary | 2. Per-Borough Results | 3. Email to Zheng Authors')
